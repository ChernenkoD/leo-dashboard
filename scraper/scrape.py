"""
Запускается в GitHub Actions по расписанию (2 раза в день).
Берёт сохранённую сессию LEO (секрет LEO_SESSION), заходит в LEO,
собирает Aufgaben и Mangelaufträge, пишет data.json.

Если сессия протухла — скрипт это обнаружит и завершится с понятной
ошибкой, ничего не поломав в data.json.
"""

import json
import re
import sys
import tempfile
import openpyxl
from datetime import datetime, timezone
from playwright.sync_api import sync_playwright

BASE = "https://leo-pro.de"
STORAGE_STATE = "storage_state.json"
OUTPUT_FILE = "../data.json"


def is_logged_out(page):
    try:
        return page.get_by_text("Alle Aufgaben").count() == 0
    except Exception:
        return True


def lines_of(text):
    return [l.strip() for l in text.split("\n") if l.strip()]


def click_next_and_wait(page):
    next_link = page.locator("text=Nächste").first
    if next_link.count() == 0:
        return False
    try:
        next_link.click()
        page.wait_for_load_state("networkidle", timeout=5000)
        return True
    except Exception:
        return False


def parse_task_block(text):
    lines = lines_of(text)
    if not lines:
        return None
    lws_match = re.search(r"(?<!M-)LWS-\d+", text)
    leg_match = re.search(r"LEG-\d+-\d+", text)
    dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", text)
    return {
        "type": lines[0],
        "address": lines[1] if len(lines) > 1 and lws_match else None,
        "lws": lws_match.group(0) if lws_match else None,
        "leg": leg_match.group(0) if leg_match else None,
        "due": dates[-1] if dates else None,
    }


def parse_tasks(page):
    page.goto(f"{BASE}/index.php")
    page.wait_for_load_state("networkidle")
    if is_logged_out(page):
        raise RuntimeError("Сессия LEO протухла — нужно перелогиниться локально и обновить LEO_SESSION secret")

    seen_keys = set()
    tasks = []
    for _ in range(20):
        blocks = page.locator("text=/(?<!M-)LWS-\\d+/").all()
        new_found = False
        for block in blocks:
            row = block.locator("xpath=ancestor::tr[1]")
            if row.count() == 0:
                continue
            text = row.first.inner_text().strip()
            task = parse_task_block(text)
            if not task:
                continue
            key = (task["lws"], task["type"], task["due"])
            if key in seen_keys:
                continue
            seen_keys.add(key)
            tasks.append(task)
            new_found = True

        if not new_found:
            break
        if not click_next_and_wait(page):
            break

    return tasks


def parse_mangel_block(text):
    lines = lines_of(text)
    id_match = re.search(r"M-LWS-\d+-\d+", text)
    if not id_match:
        return None

    status_match = re.search(r"\(([a-zA-Zäöü]+)\)", text)
    leg_match = re.search(r"LEG-\d+-\d+", text)
    fortschritt_match = re.search(r"(\d+)%\s*abgeschlossen", text)
    dates = re.findall(r"\d{2}\.\d{2}\.\d{4}", text)

    address = None
    lage = None
    bauleiter = None
    innendienst = None
    anzahl = None

    try:
        idx = lines.index("Ausführungsbeginn:")
        values = lines[idx + 4: idx + 8]
        if len(values) == 4:
            bauleiter, innendienst = values[2], values[3]
    except ValueError:
        pass

    for i, l in enumerate(lines):
        if l.startswith("Lage:"):
            lage = l.replace("Lage:", "").strip()
        if re.match(r"^[A-ZÄÖÜ].*\d.*,\s*\d{4,5}\s", l):
            address = l

    if lines and not address and len(lines) > 2:
        address = lines[2]  # лучшее приближение, если регулярка не сработала

    if lines and re.match(r"^\d+$", lines[-1]):
        anzahl = int(lines[-1])

    return {
        "id": id_match.group(0),
        "status": status_match.group(1) if status_match else None,
        "address": address,
        "lage": lage,
        "leg": leg_match.group(0) if leg_match else None,
        "fortschritt": int(fortschritt_match.group(1)) if fortschritt_match else None,
        "ausfuehrungsbeginn": dates[0] if len(dates) > 0 else None,
        "fertigstellung": dates[1] if len(dates) > 1 else None,
        "bauleiter": bauleiter,
        "innendienst": innendienst,
        "anzahl": anzahl,
    }


def parse_positionen(page):
    """
    Структура на странице детали:
      div.section → number (.inner), .lv-nummer p.top (код), p.bottom (gewerk), .badge-icon .label (статус)
    Рядом в DOM (siblings) лежат: p.kurztext (leistung), p.text-zusatz (bereich),
      div.text-apos p.kurztext (Mangelbeschreibung), div.section-price (menge)
    """
    positionen = []
    # Каждая позиция: section-description, section-text, section-price — соседи внутри apos-list-position
    containers = page.locator("div.apos-list-position").all()
    for container in containers:
        try:
            sec_desc = container.locator("div.section-description").first
            if sec_desc.count() == 0:
                continue

            code_el = sec_desc.locator("p.top").first
            if code_el.count() == 0:
                continue
            code = code_el.inner_text().strip()
            if not re.match(r"\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
                continue

            gewerk = None
            g = sec_desc.locator("p.bottom").first
            if g.count() > 0:
                gewerk = g.inner_text().strip()

            status = None
            badge_div = sec_desc.locator(".badge-icon").first
            if badge_div.count() > 0:
                # Вариант 1: текст в .label
                label = badge_div.locator(".label").first
                if label.count() > 0:
                    status = label.inner_text().strip() or None
                # Вариант 2: весь текст badge-icon
                if not status:
                    status = badge_div.inner_text().strip() or None
                # Вариант 3: title атрибут
                if not status:
                    for el in badge_div.locator("[title]").all():
                        t = (el.get_attribute("title") or "").strip()
                        if t:
                            status = t
                            break
                # Вариант 4: класс иконки → маппинг на текст
                if not status:
                    html_badge = badge_div.inner_html()
                    if "success" in html_badge or "fa-check" in html_badge:
                        status = "Mangel behoben & geprüft"
                    elif "warning" in html_badge or "fa-clock" in html_badge:
                        status = "Mangel behoben"
                    elif "fa-thumbs" in html_badge or "outline" in html_badge:
                        status = "angenommen"

            sec_text = container.locator("div.section-text").first

            leistung = None
            bereich = None
            mangel_beschreibung = None

            if sec_text.count() > 0:
                # Leistung — первый p.kurztext не содержащий "Mangelbeschreibung"
                for el in sec_text.locator("p.kurztext").all():
                    t = el.inner_text().strip()
                    if t and "Mangelbeschreibung" not in t and len(t) > 5:
                        leistung = t
                        break

                # Bereich
                z = sec_text.locator("p.text-zusatz").first
                if z.count() > 0:
                    bereich = z.inner_text().strip()

                # Mangelbeschreibung — p.kurztext после метки
                kurzs = [e.inner_text().strip() for e in sec_text.locator("p.kurztext").all()]
                for i, t in enumerate(kurzs):
                    if "Mangelbeschreibung" in t and i + 1 < len(kurzs):
                        mangel_beschreibung = kurzs[i + 1]
                        break

            menge = None
            price_el = container.locator("div.section-price").first
            if price_el.count() > 0:
                menge = price_el.inner_text().strip().replace("\n", " ")

            positionen.append({
                "code": code,
                "gewerk": gewerk,
                "status": status,
                "leistung": leistung,
                "bereich": bereich,
                "mangel_beschreibung": mangel_beschreibung,
                "menge": menge,
            })
        except Exception:
            continue

    return positionen


def collect_mangel_list(page):
    """Шаг 1: собираем все Mängel и их URL со списка, без заходов внутрь."""
    page.goto(f"{BASE}/index.php")
    page.wait_for_load_state("networkidle")
    page.click("text=Projekte")
    page.click("text=Mangelaufträge")
    page.wait_for_load_state("networkidle")

    seen_ids = set()
    items = []  # list of (mangel_dict, detail_url)

    for _ in range(20):
        blocks = page.locator("text=/M-LWS-\\d+-\\d+/").all()
        new_found = False
        for block in blocks:
            row = block.locator("xpath=ancestor::tr[1]")
            if row.count() == 0:
                continue
            text = row.first.inner_text().strip()
            m = parse_mangel_block(text)
            if not m or m["id"] in seen_ids:
                continue
            seen_ids.add(m["id"])

            # Запоминаем URL детали если есть ссылка
            detail_url = None
            link = row.first.locator("a[href*='mangel']").first
            if link.count() == 0:
                link = row.first.locator("a").first
            if link.count() > 0:
                href = link.get_attribute("href")
                if href:
                    detail_url = href if href.startswith("http") else f"{BASE}/{href.lstrip('/')}"

            m["leo_url"] = detail_url
            items.append((m, detail_url))
            new_found = True

        if not new_found:
            break
        if not click_next_and_wait(page):
            break

    return items


def parse_mangel(page):
    items = collect_mangel_list(page)
    print(f"Найдено {len(items)} Mängel в списке")

    maengel = []
    for i, (m, detail_url) in enumerate(items):
        if detail_url:
            try:
                page.goto(detail_url)
                page.wait_for_load_state("networkidle", timeout=10000)
                # Отладка: сохраняем HTML первой страницы
                if i == 0:
                    html = page.content()
                    with open("../debug_detail.html", "w", encoding="utf-8") as f:
                        f.write(html)
                    print(f"  DEBUG title={page.title()}, html={len(html)}b")
                    # Все видимые текстовые строки страницы
                    body_text = page.locator("body").inner_text()
                    print("  DEBUG body[:2000]:", body_text[:2000])
                m["positionen"] = parse_positionen(page)
                print(f"  {m['id']}: {len(m['positionen'])} позиций")
            except Exception as e:
                print(f"  {m['id']}: ошибка — {e}")
                m["positionen"] = []
        else:
            m["positionen"] = []
        maengel.append(m)

    return maengel


def parse_amount(val):
    """Excel float 18196.64 или строка '18.196,64 EUR' → float"""
    if val is None:
        return None
    # Уже число из Excel
    if isinstance(val, (int, float)):
        return float(val) if val > 0 else None
    # Строка с немецким форматом
    m = re.search(r"([\d.]+,\d{2})", str(val))
    if m:
        try:
            return float(m.group(1).replace(".", "").replace(",", "."))
        except Exception:
            pass
    return None


def fmt_date(val):
    if val is None:
        return None
    if hasattr(val, "strftime"):
        # Excel хранит "нет даты" как 0 → 1900-01-00, игнорируем
        if hasattr(val, "year") and val.year < 1970:
            return None
        return val.strftime("%d.%m.%Y")
    s = str(val).strip()
    return s if s and s != "0" else None


def download_excel(page, click_fn, label, tmpdir, timeout=120000):
    """Скачивает Excel файл, возвращает путь."""
    path = f"{tmpdir}/{label}.xlsx"
    with page.expect_download(timeout=timeout) as dl_info:
        click_fn()
    dl = dl_info.value
    dl.save_as(path)
    print(f"  Скачан {label}: {dl.suggested_filename}")
    return path


def parse_projects(page):
    """
    Два Excel-файла, ноль пагинации — быстро и надёжно:
      1. Herunterladen (Übersicht) — LWS, адрес, Lage, Bauleiter, даты, Baustopp (кол. Q/R)
      2. P-03 (Berichte)           — суммы, Fortschritt, Status
    Мерж по LWS.
    """
    with tempfile.TemporaryDirectory() as tmpdir:

        # ── 1. Herunterladen Excel с Übersicht ──────────────────────────────
        # Колонки (0-based): A=0 LWS, C=2 Straße, D=3 Nr, E=4 PLZ, F=5 Ort,
        #   H=7 Lage, I=8 Bauleiter, J=9 AusfBeginn, K=10 Fertigstellung,
        #   Q=16 Baustopp Start, R=17 Baustopp Ende
        page.goto(f"{BASE}/index.php")
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        page.click("text=Projekte")
        page.click("text=Übersicht")
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        page.wait_for_timeout(2000)  # ждём рендер таблицы

        # ── URL-кэш: листаем "Laufende Aufträge" пока на Übersicht ──────────
        url_map = {}
        try:
            with open("../project_urls.json", encoding="utf-8") as f:
                url_map = json.load(f)
            print(f"  URL-кэш загружен: {len(url_map)} записей")
        except FileNotFoundError:
            pass

        # Собираем URL только для проектов которых нет в кэше
        # DataTables: кнопка Next имеет класс .next, disabled когда последняя стр.
        pg = 0
        while pg < 200:  # hard cap
            pg += 1
            for link in page.locator("h5.section a.link-bold").all():
                try:
                    lws = link.inner_text().strip()
                    href = link.get_attribute("href") or ""
                    if lws and href and lws not in url_map:
                        url_map[lws] = href if href.startswith("http") else f"{BASE}/{href.lstrip('/')}"
                except Exception:
                    continue
            nxt = page.locator("#datatable_auftrag_laufend_paginate .paginate_button.next:not(.disabled)").first
            if nxt.count() == 0:
                break
            try:
                nxt.click()
                page.wait_for_load_state("domcontentloaded", timeout=10000)
                page.wait_for_timeout(300)
            except Exception:
                break

        with open("../project_urls.json", "w", encoding="utf-8") as f:
            json.dump(url_map, f, ensure_ascii=False, indent=2)
        print(f"  URL-кэш обновлён: {len(url_map)} записей (прошли {pg} стр.)")

        base_data = {}  # lws → dict
        try:
            xl_path = download_excel(
                page,
                lambda: page.locator("a:has-text('Herunterladen'), button:has-text('Herunterladen')").first.click(),
                "herunterladen", tmpdir, timeout=120000
            )
            wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            print(f"  Herunterladen колонки: {headers[:25]}")

            def gi(name, *fallbacks):
                """Индекс колонки по имени или фоллбэк-имени."""
                for n in (name,) + fallbacks:
                    try:
                        return headers.index(n)
                    except ValueError:
                        continue
                return None

            # Определяем индексы по именам (с фоллбэками) + хардкод по позиции
            I_LWS  = gi("Projektnummer", "LWS") or 0
            I_STR  = gi("Straße", "Strasse") or 2
            I_NR   = gi("Nr.", "Hausnummer") or 3
            I_PLZ  = gi("PLZ") or 4
            I_ORT  = gi("Ort", "Stadt") or 5
            I_LAGE = gi("Lage") or 7
            I_BL   = gi("Bauleitung AG", "Bauleiter") or 8
            I_STA  = gi("Ausführungsbeginn Plan", "Ausführungsbeginn") or 9
            I_END  = gi("Fertigstellung Plan", "Fertigstellung") or 10
            I_BSS  = gi("Baustopp Start", "Baustopp Beginn") or 16
            I_BSE  = gi("Baustopp Ende") or 17

            def g(row, idx):
                return row[idx] if idx is not None and idx < len(row) else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                lws = str(g(row, I_LWS) or "").strip()
                if not lws or not lws.startswith("LWS-"):
                    continue
                strasse = str(g(row, I_STR) or "").strip()
                nr      = str(g(row, I_NR) or "").strip()
                plz     = str(g(row, I_PLZ) or "").strip()
                ort     = str(g(row, I_ORT) or "").strip()
                address = ", ".join(filter(None, [f"{strasse} {nr}".strip(), f"{plz} {ort}".strip()]))
                bs_start = fmt_date(g(row, I_BSS))
                bs_ende  = fmt_date(g(row, I_BSE))
                base_data[lws] = {
                    "address":        address or None,
                    "lage":           str(g(row, I_LAGE) or "").strip() or None,
                    "bauleiter":      str(g(row, I_BL) or "").strip() or None,
                    "start":          fmt_date(g(row, I_STA)),
                    "ende":           fmt_date(g(row, I_END)),
                    "baustopp":       bool(bs_start),
                    "baustopp_start": bs_start,
                    "baustopp_ende":  bs_ende,
                }
            wb.close()
            print(f"  Herunterladen: {len(base_data)} строк")
        except Exception as e:
            print(f"  WARN Herunterladen ошибка: {e} — продолжаем без доп. данных")

        # ── 2. P-03 из Berichte — суммы + статус + fortschritt ─────────────
        page.goto(f"{BASE}/index.php")
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        page.click("text=Berichte")
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        page.wait_for_timeout(1000)

        p03_path = download_excel(
            page,
            lambda: page.locator("tr", has_text="P-03").locator("a, button").first.click(),
            "p03", tmpdir, timeout=60000
        )
        wb = openpyxl.load_workbook(p03_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  P-03 колонки: {headers[:25]}")

        def gi2(name, *fallbacks):
            for n in (name,) + fallbacks:
                try:
                    return headers.index(n)
                except ValueError:
                    continue
            return None

        # P-03: LWS в колонке A (Projektnummer), суммы, статус, fortschritt
        I2_LWS  = gi2("Projektnummer") or 0
        I2_AMT  = gi2("Auftragsvolumen gesamt", "Auftragsvolumen")
        I2_FRT  = gi2("Projektfortschritt", "Fortschritt")
        I2_STA  = gi2("Status")
        # Если Herunterladen не дал адрес — берём из P-03
        I2_STR  = gi2("Straße", "Strasse") or 2
        I2_NR   = gi2("Nr.") or 3
        I2_PLZ  = gi2("PLZ") or 4
        I2_ORT  = gi2("Ort") or 5
        I2_LAGE = gi2("Lage") or 7
        I2_BL   = gi2("Bauleitung AG", "Bauleiter") or 8
        I2_STRT = gi2("Ausführungsbeginn Plan", "Ausführungsbeginn") or 9
        I2_END  = gi2("Fertigstellung Plan", "Fertigstellung") or 10

        CLOSED = ["beendet", "abgeschlossen", "schlussgerechnet", "storniert"]

        projects = []
        seen_lws = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            lws = str(row[I2_LWS] if I2_LWS < len(row) else "").strip()
            if not lws or not lws.startswith("LWS-"):
                continue
            if lws in seen_lws:
                continue  # дубликат — пропускаем
            seen_lws.add(lws)

            status = str(row[I2_STA] if I2_STA is not None and I2_STA < len(row) else "").strip()
            fortschritt = 0
            try:
                raw = row[I2_FRT] if I2_FRT is not None and I2_FRT < len(row) else None
                if raw is not None:
                    v = float(str(raw).replace("%", "").strip())
                    fortschritt = int(v * 100) if v <= 1.0 else int(v)
            except Exception:
                pass

            abgeschlossen = any(k in status.lower() for k in CLOSED) or fortschritt >= 100

            base = base_data.get(lws, {})

            def p03_val(idx):
                return row[idx] if idx is not None and idx < len(row) else None

            # Берём адрес из Herunterladen если есть, иначе из P-03
            address = base.get("address")
            if not address:
                strasse = str(p03_val(I2_STR) or "").strip()
                nr      = str(p03_val(I2_NR) or "").strip()
                plz     = str(p03_val(I2_PLZ) or "").strip()
                ort     = str(p03_val(I2_ORT) or "").strip()
                address = ", ".join(filter(None, [f"{strasse} {nr}".strip(), f"{plz} {ort}".strip()])) or None

            projects.append({
                "lws":            lws,
                "address":        address,
                "lage":           base.get("lage") or str(p03_val(I2_LAGE) or "").strip() or None,
                "bauleiter":      base.get("bauleiter") or str(p03_val(I2_BL) or "").strip() or None,
                "start":          base.get("start") or fmt_date(p03_val(I2_STRT)),
                "ende":           base.get("ende") or fmt_date(p03_val(I2_END)),
                "amount":         parse_amount(p03_val(I2_AMT)),
                "fortschritt":    fortschritt,
                "status":         status or None,
                "abgeschlossen":  abgeschlossen,
                "baustopp":       base.get("baustopp", False),
                "baustopp_start": base.get("baustopp_start"),
                "baustopp_ende":  base.get("baustopp_ende"),
                "leo_url":        url_map.get(lws),
            })

        wb.close()

    with_amount   = sum(1 for p in projects if p["amount"])
    abgeschl      = sum(1 for p in projects if p["abgeschlossen"])
    with_baustopp = sum(1 for p in projects if p["baustopp"])
    print(f"Итого: {len(projects)} проектов | суммы: {with_amount} | закрытых: {abgeschl} | Baustopp: {with_baustopp}")
    return projects


def main():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(storage_state=STORAGE_STATE)
        page = context.new_page()

        try:
            tasks = parse_tasks(page)
            maengel = parse_mangel(page)
            projects = parse_projects(page)
        except RuntimeError as e:
            print(f"ОШИБКА: {e}", file=sys.stderr)
            sys.exit(1)

        # Помечаем проекты у которых есть Mängel (по базовому LWS номеру)
        # M-LWS-82670-2 → LWS-82670, проект LWS-82670 → совпадение
        mangel_lws_set = set()
        for m in maengel:
            match = re.search(r"LWS-\d+", m.get("id", ""))
            if match:
                mangel_lws_set.add(match.group(0))

        for p in projects:
            match = re.search(r"LWS-\d+", p.get("lws", ""))
            p["has_mangel"] = bool(match and match.group(0) in mangel_lws_set)

        data = {
            "updatedAt": datetime.now(timezone.utc).isoformat(),
            "tasks": tasks,
            "maengel": maengel,
            "projects": projects,
        }

        with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"Сохранено {len(tasks)} задач, {len(maengel)} Mängel, {len(projects)} проектов в {OUTPUT_FILE}")
        browser.close()


if __name__ == "__main__":
    main()
